import openpyxl,numpy
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill
myfont=Font(bold=True)
highlight = PatternFill(start_color='FFA533',
                   end_color='FFA533',
                   fill_type='solid')
n=0;changeshift=0;free=[0]*31; gen=[0]*31; mor=[0]*31; eve=[0]*31

wb=openpyxl.load_workbook('ROASTER_INPUT.xlsx')
sheet=wb.active
totaldays=sheet['B1'].value
sfirstday=sheet['B2'].value

if("SUN"==sfirstday): firstday=1
elif("MON"==sfirstday): firstday=2
elif("TUE"==sfirstday): firstday=3 
elif("WED"==sfirstday): firstday=4
elif("THU"==sfirstday): firstday=5  
elif("FRI"==sfirstday): firstday=6  
elif("SAT"==sfirstday): firstday=7

weekday=firstday

wb1 = openpyxl.Workbook() #new workbook
sheet1 = wb1.active
for i in range(totaldays):                              #writing dates and weekdays in excel 1st and 2nd row
	sheet1.cell(row=1,column=i+2).value=i+1
	sheet1.cell(row=1,column=i+2).font=sheet1.cell(row=2,column=i+2).font=myfont
	sheet1.column_dimensions[get_column_letter(i+2)].width=5
	if weekday>7: weekday=1
	if weekday==1: sheet1.cell(row=2,column=i+2).value='SUN'
	elif weekday==2: sheet1.cell(row=2,column=i+2).value='MON'
	elif weekday==3: sheet1.cell(row=2,column=i+2).value='TUE'
	elif weekday==4: sheet1.cell(row=2,column=i+2).value='WED'
	elif weekday==5: sheet1.cell(row=2,column=i+2).value='THU'
	elif weekday==6: sheet1.cell(row=2,column=i+2).value='FRI'
	elif weekday==7: sheet1.cell(row=2,column=i+2).value='SAT'
	weekday=weekday+1 

GME_employees=sheet['B3'].value
G_employees=sheet['G3'].value
rotatable=numpy.empty(((GME_employees+G_employees,totaldays)),dtype=str)
rotatable.fill('F ')
employeename=['']*(GME_employees+G_employees)
prefshift=['']*(GME_employees+G_employees)

for i in range(GME_employees+G_employees):
	if i<GME_employees:
		employeename[i]=sheet.cell(row=5+i,column=1).value
		srest=sheet.cell(row=5+i,column=2).value
		altoff=sheet.cell(row=5+i,column=3).value
		prefshift[i]=sheet.cell(row=5+i,column=4).value
	else:
		employeename[i]=sheet.cell(row=5+i-GME_employees ,column=6).value
		srest=sheet.cell(row=5+i-GME_employees,column=7).value
		altoff=sheet.cell(row=5+i-GME_employees,column=8).value
		prefshift[i]='G'
	if("SUN"==srest): rest=1
	elif("MON"==srest): rest=2
	elif("TUE"==srest): rest=3
	elif("WED"==srest): rest=4
	elif("THU"==srest): rest=5
	elif("FRI"==srest): rest=6
	elif("SAT"==srest): rest=7
	while(1):
		date=1+7*n+rest-firstday
		if(date>totaldays): break
		if(date>0):
			rotatable[i][date-1]='R'
		n+=1
	n=0
	if((rest-1)<firstday):
			if(altoff==1):
				date=1+7+(rest-1)-firstday
				rotatable[i][date-1]='A'
				date=1+7*3+(rest-1)-firstday;
				rotatable[i][date-1]='A'   
			elif(altoff==2):
				 date=1+7*2+(rest-1)-firstday
				 rotatable[i][date-1]='A'
				 date=1+7*4+(rest-1)-firstday
				 rotatable[i][date-1]='A'
	else:
			if(altoff==1):
				date=1+(rest-1)-firstday
				rotatable[i][date-1]='A'
				date=1+7*2+(rest-1)-firstday
				rotatable[i][date-1]='A'
			elif(altoff==2):
				date=1+7*1+(rest-1)-firstday
				rotatable[i][date-1]='A'
				date=1+7*3+(rest-1)-firstday
				rotatable[i][date-1]='A'

for j in range(GME_employees):
	for k in range(totaldays):
		if(rotatable[j][k]=='F'): free[k]+=1
for i in range(GME_employees):
	shiftvar=prefshift[i]
	for j in range(totaldays):
		if(rotatable[i][j]=='F'):
			changeshift=1
			if(shiftvar=='G'):
							 if((free[j]+mor[j]+eve[j])>4): 
							 	rotatable[i][j]='G'; gen[j]+=1; free[j]-=1    
							 elif(mor[j]<2):  
							 	rotatable[i][j]='M'; mor[j]+=1; free[j]-=1
							 else: 
							 	rotatable[i][j]='E'; eve[j]+=1; free[j]-=1
			elif(shiftvar=='M'): 
							 if(mor[j]<2): 
							 	rotatable[i][j]='M'; mor[j]+=1; free[j]-=1
							 elif(eve[j]<2):  
							 	rotatable[i][j]='E'; eve[j]+=1; free[j]-=1;  
							 else: 
							 	rotatable[i][j]='G'; gen[j]+=1; free[j]-=1
			elif(shiftvar=='E'):	
							 if(eve[j]<2): 
							 	rotatable[i][j]='E'; eve[j]+=1; free[j]-=1; 
							 elif((free[j]+mor[j]+eve[j])>4): 
							 	rotatable[i][j]='G'; gen[j]+=1; free[j]-=1   
							 else: 
							 	rotatable[i][j]='M'; mor[j]+=1; free[j]-=1
		elif(rotatable[i][j]=='R'): 
			if(changeshift==1):
				if(shiftvar=='G'): shiftvar='M'
				elif(shiftvar=='M'): shiftvar='E'
				elif(shiftvar=='E'): shiftvar='G'
	changeshift=0
for i in range(GME_employees,GME_employees+G_employees):
	for j in range(totaldays):
		if(rotatable[i][j]=='F'): rotatable[i][j]='G'

for i in range(GME_employees):
	sheet1.cell(row=3+i,column=1).font=myfont
	sheet1.cell(row=3+i,column=1).value=employeename[i]
	for j in range(totaldays): 
		if rotatable[i][j]=='A' or rotatable[i][j]=='R':
			sheet1.cell(row=3+i,column=2+j).fill=highlight
		sheet1.cell(row=3+i,column=2+j).value=rotatable[i][j]	
for i in range(GME_employees,GME_employees+G_employees):
	sheet1.cell(row=3+i+1,column=1).font=myfont
	sheet1.cell(row=3+i+1,column=1).value=employeename[i]
	for j in range(totaldays):
		if rotatable[i][j]=='A' or rotatable[i][j]=='R':
			sheet1.cell(row=3+i+1,column=2+j).fill=highlight
		sheet1.cell(row=3+i+1,column=2+j).value=rotatable[i][j]
wb1.save('NM_ROASTER_PLANNED_BY_SATISH.xlsx')
#t=input('Developer: satishkushwah50@gmail.com\n\nRoaster created succesfully, press enter to exit')